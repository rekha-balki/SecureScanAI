"""
Findings export: JSON and Excel (FRS Section 11 lists Executive PDF,
Technical PDF, Excel, JSON, and HTML as required report formats; this
module covers Excel and JSON. PDF is report_service.py).
"""

from __future__ import annotations

import io
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.domains.finding.domain.aggregates.finding import Finding
from app.domains.scan.domain.aggregates.scan import Scan

_SEVERITY_FILL = {
    "critical": "F0475C",
    "high": "FF8A3D",
    "medium": "F5C242",
    "low": "4E9CFF",
    "informational": "9AA5B1",
}

_HEADER_FILL = "0F172A"


def generate_findings_json(scan: Scan, findings: list[Finding]) -> bytes:
    payload = {
        "scan": {
            "id": scan.id,
            "name": scan.name,
            "target_url": scan.target_url,
            "status": scan.status.value,
            "pages_crawled": scan.pages_crawled,
            "findings_count": len(findings),
            "started_at": scan.started_at.isoformat() if scan.started_at else None,
            "completed_at": scan.completed_at.isoformat() if scan.completed_at else None,
        },
        "findings": [
            {
                "id": f.id,
                "plugin_id": f.plugin_id,
                "plugin_name": f.plugin_name,
                "category": f.category,
                "severity": f.severity.value,
                "confidence": f.confidence.value,
                "affected_url": f.affected_url,
                "description": f.description,
                "evidence": f.evidence,
                "recommendation": f.recommendation,
                "business_impact": f.business_impact,
                "technical_impact": f.technical_impact,
                "owasp_reference": f.owasp_reference,
                "cwe_reference": f.cwe_reference,
                "cvss_score": f.cvss_score,
                "cvss_vector": f.cvss_vector,
                "status": f.status.value,
                "assigned_user_id": f.assigned_user_id,
                "created_at": f.created_at.isoformat(),
            }
            for f in findings
        ],
        "exported_at": datetime.utcnow().isoformat() + "Z",
    }

    return json.dumps(payload, indent=2).encode("utf-8")


def generate_findings_excel(scan: Scan, findings: list[Finding]) -> bytes:
    workbook = Workbook()

    _build_summary_sheet(workbook, scan, findings)
    _build_findings_sheet(workbook, findings)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _build_summary_sheet(workbook: Workbook, scan: Scan, findings: list[Finding]) -> None:
    sheet = workbook.active
    sheet.title = "Summary"

    sheet["A1"] = "SecureScan AI - Scan Summary"
    sheet["A1"].font = Font(size=14, bold=True)

    rows = [
        ("Scan Name", scan.name),
        ("Target URL", scan.target_url),
        ("Status", scan.status.value),
        ("Pages Crawled", scan.pages_crawled),
        ("Total Findings", len(findings)),
        ("Started At", scan.started_at.isoformat() if scan.started_at else "-"),
        ("Completed At", scan.completed_at.isoformat() if scan.completed_at else "-"),
    ]

    for i, (label, value) in enumerate(rows, start=3):
        sheet.cell(row=i, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=i, column=2, value=str(value))

    severity_counts: dict[str, int] = {}
    for f in findings:
        severity_counts[f.severity.value] = severity_counts.get(f.severity.value, 0) + 1

    start_row = len(rows) + 5
    sheet.cell(row=start_row, column=1, value="Severity Breakdown").font = Font(bold=True)

    for i, severity in enumerate(("critical", "high", "medium", "low", "informational")):
        row = start_row + 1 + i
        sheet.cell(row=row, column=1, value=severity.title())
        count_cell = sheet.cell(row=row, column=2, value=severity_counts.get(severity, 0))
        fill = _SEVERITY_FILL.get(severity)
        if fill:
            count_cell.fill = PatternFill("solid", fgColor=fill)

    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 60


def _build_findings_sheet(workbook: Workbook, findings: list[Finding]) -> None:
    sheet = workbook.create_sheet("Findings")

    headers = [
        "Severity",
        "Confidence",
        "CVSS (est.)",
        "Category",
        "Plugin",
        "Affected URL",
        "Description",
        "Evidence",
        "Recommendation",
        "OWASP",
        "CWE",
        "Status",
        "Created At",
    ]

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
        cell.alignment = Alignment(vertical="center")

    ordered = sorted(
        findings,
        key=lambda f: ("critical", "high", "medium", "low", "informational").index(
            f.severity.value
        ),
    )

    for row_idx, finding in enumerate(ordered, start=2):
        values = [
            finding.severity.value.title(),
            finding.confidence.value.title(),
            finding.cvss_score,
            finding.category,
            finding.plugin_name,
            finding.affected_url,
            finding.description,
            finding.evidence,
            finding.recommendation,
            finding.owasp_reference or "",
            finding.cwe_reference or "",
            finding.status.value,
            finding.created_at.isoformat(),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        severity_fill = _SEVERITY_FILL.get(finding.severity.value)
        if severity_fill:
            sheet.cell(row=row_idx, column=1).fill = PatternFill(
                "solid", fgColor=severity_fill
            )

    widths = [12, 12, 10, 20, 22, 40, 45, 45, 45, 22, 10, 12, 20]
    for i, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    sheet.freeze_panes = "A2"
